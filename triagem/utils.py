import re
import unicodedata
from datetime import datetime

from openpyxl import load_workbook

HEADER_ALIASES = {
    "codfilial": ["codfilial", "cod_filial"],
    "numped": ["numped", "num_ped", "numero_pedido"],
    "posicao": ["posicao", "posição"],
    "cod_rca": ["cod_rca", "codrca"],
    "vendedor": ["vendedor"],
    "cod_sup": ["cod_sup", "codsup"],
    "nome_supervisor": ["nome_supervisor", "supervisor", "nome_sup", "nomesupervisor"],
    "codcli": ["codcli", "cod_cli"],
    "fantasia": ["fantasia"],
    "data_pedido": ["data_pedido", "datapedido"],
    "hora_pedido": ["hora_pedido", "horapedido"],
    "vlatend": ["vlatend", "valor_atend", "vl_atend"],
    "codcob": ["codcob", "cod_cob"],
    "obs": ["obs"],
    "obs1": ["obs1", "observacao1"],
    "obs2": ["obs2", "observacao2"],
}


def normalizar_cabecalho(h):
    h = str(h or "")
    h = unicodedata.normalize("NFD", h).encode("ascii", "ignore").decode("utf-8")
    h = h.lower().strip()
    h = re.sub(r"[^a-z0-9]+", "_", h).strip("_")
    return h


def mapear_cabecalhos(cabecalhos_brutos):
    mapa = {}
    for bruto in cabecalhos_brutos:
        norm = normalizar_cabecalho(bruto)
        for campo, aliases in HEADER_ALIASES.items():
            if norm in aliases:
                mapa[bruto] = campo
    return mapa


def ler_planilha(arquivo):
    """Lê um arquivo .xlsx (InMemoryUploadedFile ou caminho) e retorna
    uma lista de dicts com as colunas já mapeadas para os nomes internos."""
    wb = load_workbook(arquivo, data_only=True)
    ws = wb.active
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        return []

    cabecalhos = linhas[0]
    mapa = mapear_cabecalhos(cabecalhos)

    resultado = []
    for linha in linhas[1:]:
        row = {}
        for i, valor in enumerate(linha):
            if i >= len(cabecalhos):
                continue
            campo = mapa.get(cabecalhos[i])
            if campo:
                row[campo] = valor
        if any(v not in (None, "") for v in row.values()):
            resultado.append(row)
    return resultado


def gerar_chave(row):
    """A chave real do pedido é filial + número do pedido (NUMPED).
    POSICAO NÃO serve como chave: é um status (F/L/B/C), se repete em
    milhares de linhas diferentes."""
    codfilial = row.get("codfilial")
    numped = row.get("numped")
    if codfilial not in (None, "") and numped not in (None, ""):
        return f"ped_{str(codfilial).strip()}_{str(numped).strip()}"

    posicao = row.get("posicao")
    if posicao not in (None, ""):
        return f"pos_{str(posicao).strip()}_{row.get('codcli')}_{row.get('data_pedido')}"
    return f"k_{row.get('codcli')}_{row.get('data_pedido')}_{row.get('hora_pedido')}"


def parse_data(valor):
    if valor in (None, ""):
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if hasattr(valor, "year") and hasattr(valor, "month"):  # já é date
        return valor
    if isinstance(valor, str):
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(valor.strip(), fmt).date()
            except ValueError:
                continue
    return None


PALAVRAS_RECOMPOSICAO = [
    "recomposicao", "recomposição", "recompor", "recompondo",
    "recompos", "recomp",
]

PALAVRAS_PRAZO = [
    "prazo", "prazos", "fora do prazo", "sem prazo",
    "prorrogacao", "prorrogação", "prorrogar",
]

PALAVRAS_BONIFICACAO = [
    "bonificacao", "bonificação", "bonific", "bonif",
]

PALAVRAS_INFRAECOMMERCE = [
    "infraecommerce", "infracommerce",       # com e sem o "e"
    "infra ecommerce", "infra e-commerce", "infra e commerce",
    "rps",
]

# sequências numéricas que representam prazo mesmo sem a palavra "prazo"
# ex: "7 14 21 28"  |  "14/21/28/35/42/49/56"  |  "de7 a 35"
PADRAO_SEQUENCIA_NUMEROS = re.compile(r"\d+(?:[\s/,]+\d+){1,}")
PADRAO_INTERVALO = re.compile(r"\bde\s?\d+\s*a\s*\d+\b", re.IGNORECASE)


def _contem_alguma(texto, palavras):
    """Verifica se o texto contém alguma das palavras/expressões, respeitando
    limite de palavra (evita falso positivo, ex: 'rps' dentro de outra palavra)."""
    for p in palavras:
        padrao = r"\b" + re.escape(p) + r"\b"
        if re.search(padrao, texto):
            return True
    return False


def _contem_prazo_numerico(texto):
    return bool(PADRAO_SEQUENCIA_NUMEROS.search(texto) or PADRAO_INTERVALO.search(texto))


def _texto_completo(row):
    """Junta as 3 colunas de observação (obs, obs1, obs2) num texto só."""
    return f"{row.get('obs') or ''} {row.get('obs1') or ''} {row.get('obs2') or ''}".lower()


def classificar_heuristico(obs_texto):
    t = (obs_texto or "").lower()

    if not t.strip():
        return "tratar", "sem_obs"
    if _contem_alguma(t, PALAVRAS_RECOMPOSICAO):
        return "recomposicao", "palavra_recomposicao"
    if _contem_alguma(t, PALAVRAS_PRAZO) or _contem_prazo_numerico(t):
        return "prazo", "palavra_ou_numero_prazo"
    return "tratar", "sem_padrao_reconhecido"


def categorizar(row, suporte_infra_nome=None):
    """Aplica a regra de negócio e devolve (categoria, motivo, sugestao_atribuicao)."""
    codcob = str(row.get("codcob") or "").strip().lower()
    texto = _texto_completo(row)

    # 1) Nívea: qualquer pedido integrado ao Infracommerce/RPS
    if _contem_alguma(texto, PALAVRAS_INFRAECOMMERCE):
        return "nivea", "Integrado Infracommerce (RPS)", suporte_infra_nome

    # 2) Bonificação: via codcob=BNF OU via palavra nas obs (obs/obs1/obs2)
    if codcob == "bnf":
        return "bonificacao", "codcob = BNF (Bonificação)", None
    if _contem_alguma(texto, PALAVRAS_BONIFICACAO):
        return "bonificacao", "Bonificação identificada na obs", None

    if codcob == "bk":
        return "liberar", "codcob = BK", None
    if codcob == "dh":
        return "liberar", "codcob = DH", None
    if codcob == "dep":
        categoria, motivo_tecnico = classificar_heuristico(texto)
        motivos = {
            "sem_obs": "codcob = DEP — obs/obs1/obs2 vazias, triagem manual necessária",
            "palavra_recomposicao": "codcob = DEP — recomposição identificada na obs",
            "palavra_ou_numero_prazo": "codcob = DEP — prazo identificado na obs",
            "sem_padrao_reconhecido": "codcob = DEP — sem padrão reconhecido na obs, triagem manual necessária",
        }
        return categoria, motivos[motivo_tecnico], None

    return "sem_categoria", "codcob não mapeado", None
